from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
import pytest
import openpyxl
import global_constants
from pathlib import Path
from datetime import date

class Test_Demo_Class:

    #her testten once
    def setup_method(self):
        options = Options()
        options.page_load_strategy = 'normal'
        chrome_driver = ChromeDriverManager().install()
        self.driver = webdriver.Chrome(chrome_driver, options=options)
        self.driver.get("https://www.saucedemo.com")
        self.driver.maximize_window()

        #gunun tarihi
        self.folder_path = str(date.today()) #24.03.2023 

        #klasoru olustur
        Path(self.folder_path).mkdir(exist_ok=True)

    #dekoraterden cagirilacak fonksiyon icin self yazmiyorum.
    def get_data():
        #veriyi al

        excel_file = openpyxl.load_workbook("data/invalid_login_test.xlsx")
        selected_sheet = excel_file["Sheet1"]
        total_rows= selected_sheet.max_row
        data =[]
        for i in range (2,total_rows+1):
            username = selected_sheet.cell(i,1).value
            password = selected_sheet.cell(i,2).value
            tupple_data = (username,password)
            data.append(tupple_data)
        return data

    #her testten sonra
    def teardown_method(self):
        self.driver.quit()

    @pytest.mark.parametrize("username,password" , get_data())
    def test_valid_login(self,username: Any, password: Any):  
        # WebDriverWait(driver=self.driver, timeout=5).until(ec.visibility_of_element_located((By.ID,"user-name")))
        self.wait_for_element_visible((By.ID,"user-name"),3)
        username_input = self.driver.find_element(By.ID,"user-name")
        password_input = self.driver.find_element(By.ID,"password")
      
       
        #ACTION CHAINS
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(username_input, username)
        actions.send_keys_to_element(password_input, password)
        actions.perform()
        login_click = self.driver.find_element(By.ID,"login-button").click()

        #JAVASCRIPT -> execute_script(JS comes here)
        # self.driver.execute_script('window.scrollTo(0,500)')
        self.driver.save_screenshot(f"{self.folder_path}/test-{username}-{password}.png")
        message = self.driver.find_element(By.XPATH, '//*[@id="login_button_container"]/div/form/div[3]/h3')
        #magic string --> dont
        assert message.text == "Epic sadface: Username and password do not match any user in this service"
        
    @pytest.mark.skip()
    def test_invalid_login(self):
        self.wait_for_element_visible((By.ID,"user-name"))
        #WebDriverWait(driver=self.driver, timeout=5).until(ec.visibility_of_element_located((By.ID,"user-name")))
        username_input = self.driver.find_element(By.ID,"user-name").send_keys("1")
        # WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"password")))
        self.wait_for_element_visible((By.ID,"password"))
        password_input = self.driver.find_element(By.ID,"password").send_keys("1")
        login_click = self.driver.find_element(By.ID,"login-button").click()
        
        message = self.driver.find_element(By.XPATH, '//*[@id="login_button_container"]/div/form/div[3]/h3')
        assert message.text == "Epic sadface: Username and password do not match any user in this service"
        
    def wait_for_element_visible(self,locator,timeout=5):
        WebDriverWait(driver=self.driver, timeout=timeout).until(ec.visibility_of_element_located((locator)))

